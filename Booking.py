#run cmd 'pip install pandas openpyxl' for first time using pandas
import pandas as pd 
from openpyxl import load_workbook

print("Hello user, please find the barber availabilities: \n")

#file path to excel file
file_path = r"C:\Users\DEBRO\Desktop\New folder\BTM495\Barbershop-schedule.xlsx"

df = pd.read_excel(file_path)
workbook = load_workbook(file_path)
sheet = workbook.active

#cleaning excel output in terminal
df.columns = ["" if "Unnamed" in col else col for col in df.columns]
df = df.fillna(" ")

print(df)

# Display the menu for day selection
print("\nPlease enter the number corresponding to the day you want to book:")

# List of days corresponding to numbers 1 through 7
days_of_week = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# Print the days with their corresponding numbers
for i, day in enumerate(days_of_week, start=1):
    print(f"{i} - {day}")

# Loop until a valid input is provided
while True:
    try:
        day_number = int(input("Enter a number (1-7): "))
        
        # Check if the input is within the valid range
        if 1 <= day_number <= 7:
            selected_day = days_of_week[day_number - 1] 
            print(f"{selected_day} selected.")
            break
        else:
            print("Error: Please enter a number between 1 and 7.")
    
    #if user enters something that is not an int:
    except ValueError:
        print("Error: Invalid input. Please enter a numeric value.")



# Display the menu for barber selection
print("\nPlease enter the number corresponding to the barber you want to book:")

# List of barbers corresponding to numbers 1 through 4
barbers = ["Gonz", "Roland", "Junior", "Mahir"]

# Print the barbers with their corresponding numbers
for i, barber in enumerate(barbers, start=1):
    print(f"{i} - {barber}")

# Loop until a valid input is provided
while True:
    try:
        # Get user input
        barber_number = int(input("Enter a number (1-4): "))
        
        # Check if the input is within the valid range
        if 1 <= barber_number <= 4:
            selected_barber = barbers[barber_number - 1]  # Adjust index to match user input
            print(f"{selected_barber} selected.")
            break
        else:
            print("Error: Please enter a number between 1 and 4.")
    
    except ValueError:
        print("Error: Invalid input. Please enter a numeric value.")


#find column index of user-selected day and barber for excel sheet:
column_value = 1 + (day_number-1)*4 + barber_number

time_values = []
# Print the entire column based on the calculated column_value 
print(f"""Here are the openings for {selected_barber} on {selected_day}.
Please enter the number corresponding to the time slot you want to book: """)
for row in range(3, sheet.max_row + 1):  # Start from row 3 to skip headers
    row_value = row-2
    cell_value = sheet.cell(row=row, column=column_value).value
    if cell_value is None:
        cell_value = "Available"
    time_value = sheet.cell(row=row, column=1).value
    time_values.append(time_value)
    print(f"{row_value} - {time_value} {cell_value}")


while True:
    try:
        time_selection = int(input(f"Enter a number (1 - {len(time_values)}): "))
        
        # Check if the input is within the valid range
        if 1 <= time_selection <= len(time_values):
            row_value = time_selection+2
            cell_value = sheet.cell(row=row_value, column=column_value).value
            if cell_value is None:
                print(f"Your booking has been successful!")
                sheet.cell(row=row_value, column=column_value).value = "Booked"
                selected_time = sheet.cell(row=row_value, column=1).value
                workbook.save(file_path)
            else:
                print(f"This time slot is unavailable. Please pick a different time slot")
                continue
            break
        else:
            print(f"Error: Please enter a number between 1 and {len(time_values)}.")
    
    #if user enters something that is not an int:
    except ValueError:
        print("Error: Invalid input. Please enter a numeric value.")

#type of cut, cost, points rewarded
cut_types = [ 
    ("Haircut", 40, 4),
    ("Haircut + Beard", 50, 5),
    ("Braids", 100, 10)
]

print(f"""Here are the haircut types.
Please enter the number corresponding to the haircut type you want to book: """)

for index, (cut_type, cost, points) in enumerate(cut_types, start=1):
    print(f"{index} - Cut type: {cut_type}, Cost: ${cost}, Points: {points} rewarded")

while True:
    try:
        cut_selection = int(input(f"Enter a number (1 - {len(cut_types)}): "))

        if 1 <= cut_selection <= len(cut_types):
            selected_cut = cut_types[cut_selection-1][0]
            print(f"{selected_cut} chosen")
            break
        else:
            print(f"Error: Please enter a number between 1 and {len(cut_types)}: ")

    #if user enters something that is not an int:
    except ValueError:
        print("Error: Invalid input. Please enter a numeric value.")

fname = input ("Please enter your first name: ")
lname = input("Please enter your last name: ")
email = input("Please enter your email address: ")

print(f"""Thank you for booking your appointment with us, {fname} {lname}!
Your reservation for {selected_day} with {selected_barber} at {selected_time} is confirmed.
Your type of haircut: {selected_cut}
A confirmation email will be sent to {email}""")